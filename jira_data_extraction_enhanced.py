"""
Enhanced Jira Data Extraction with Quality Scoring
Fetch from Jira + LLM Score + Quality Metrics + JSON Export
Processes each epic end-to-end and generates comprehensive quality reports.
Designed for 100+ epics. Outputs to JSON for story generation workflows.
"""

import os
import json
import logging
import time
import warnings
import sys
import queue
import threading
from concurrent.futures import TimeoutError as FutureTimeoutError
from concurrent.futures import ThreadPoolExecutor, as_completed

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)

import requests
from requests.auth import HTTPBasicAuth
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from google.cloud import bigquery

import google.cloud.aiplatform as aiplatform
from langchain_google_vertexai import VertexAI

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def configure_logging():
    """Log to both terminal and text file."""
    log_file = os.getenv("RUN_LOG_FILE", "run.log")
    level_name = os.getenv("LOG_LEVEL", "INFO").upper()
    level = getattr(logging, level_name, logging.INFO)

    logger.setLevel(level)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(formatter)
    stream_handler.setLevel(level)
    logger.addHandler(stream_handler)

    file_handler = logging.FileHandler(log_file, mode="w", encoding="utf-8")
    file_handler.setFormatter(formatter)
    file_handler.setLevel(level)
    logger.addHandler(file_handler)

    logger.propagate = False
    logger.info(f"Logging to terminal and file: {log_file}")


configure_logging()

# ═══════════════════════════════════════════════════════════════════════════════
# Configuration
# ═══════════════════════════════════════════════════════════════════════════════
JIRA_URL = os.getenv("JIRA_BASE_URL", "")
JIRA_EMAIL = os.getenv("JIRA_EMAIL", "")
JIRA_API_TOKEN = os.getenv("JIRA_API_TOKEN", "")

# ── Epic input configuration ─────────────────────────────────────────────────
# Priority:
# 1) EPIC_KEYS_CSV env var: "CTLEP-1,CTLEP-2"
# 2) EPIC_KEYS_FILE env var: path to file, one epic key per line
# 3) epic_keys.txt in the script directory
# 4) DEFAULT_EPIC_KEYS below (empty — use epic_keys.txt instead)
DEFAULT_EPIC_KEYS = []

Q3_FIX_VERSION_PATTERNS = ["Q3", "2026 Q3", "q3", "Q326"]
STORY_POINTS_FIELDS = ["customfield_10028", "customfield_10004"]
TEAM_FIELD = "customfield_10001"

# Fix versions that mark a feature as Q3-scoped at the planning level
PLANNING_SET = {
    "CLD PC42 May06", "CLD PC43 May13", "CLD PC44 Jun10", "CLD PC45 Jul08",
    "CPPE PC42 May06", "CPPE PC43 May13", "CPPE PC44 Jun10", "CPPE PC45 Jul08",
    "CYSEC PC42 May06", "CYSEC PC43 May13", "CYSEC PC44 Jun10", "CYSEC PC45 Jul08",
    "DEA PC42 May06", "DEA PC43 May13", "DEA PC44 Jun10", "DEA PC45 Jul08",
    "DPE PC42 May06", "DPE PC43 May13", "DPE PC44 Jun10", "DPE PC45 Jul08",
    "Edge PC42 May06", "Edge PC43 May13", "Edge PC44 Jun10", "Edge PC45 Jul08",
    "EDP PC42 May06", "EDP PC43 May13", "EDP PC44 Jun10", "EDP PC45 Jul08",
    "FS PC42 May06", "FS PC43 May13", "FS PC44 Jun10", "FS PC45 Jul08",
    "GNTS PC42 May06", "GNTS PC43 May13", "GNTS PC44 Jun10", "GNTS PC45 Jul08",
    "LD PC42 May06", "LD PC43 May13", "LD PC44 Jun10", "LD PC45 Jul08",
    "PS PC42 May06", "PS PC43 May13", "PS PC44 Jun10", "PS PC45 Jul08",
    "SE PC42 May06", "SE PC43 May13", "SE PC44 Jun10", "SE PC45 Jul08",
    "Voice PC42 May06", "Voice PC43 May13", "Voice PC44 Jun10", "Voice PC45 Jul08",
    "MEDIA PC42 May06", "MEDIA PC43 May13", "MEDIA PC44 Jun10", "MEDIA PC45 Jul08",
    "Vyvx PC42 May06", "Vyvx PC43 May13", "Vyvx PC44 Jun10", "Vyvx PC45 Jul08",
}

OUTPUT_FILE_EXCEL = "Q3_2026_All_Epics_Quality_Report.xlsx"
ENABLE_BIGQUERY_UPLOAD = os.getenv("ENABLE_BIGQUERY_UPLOAD", "true").strip().lower() in {"1", "true", "yes", "y"}
SAVE_PER_EPIC_EXCEL = os.getenv("SAVE_PER_EPIC_EXCEL", "false").strip().lower() in {"1", "true", "yes", "y"}
SAVE_COMBINED_EXCEL = os.getenv("SAVE_COMBINED_EXCEL", "false").strip().lower() in {"1", "true", "yes", "y"}
BQ_APPEND_PER_EPIC = os.getenv("BQ_APPEND_PER_EPIC", "true").strip().lower() in {"1", "true", "yes", "y"}
BQ_CLEAR_TABLE_BEFORE_RUN = os.getenv("BQ_CLEAR_TABLE_BEFORE_RUN", "true").strip().lower() in {"1", "true", "yes", "y"}
ALL_EPICS_BIGQUERY_TABLE_ID = os.getenv("ALL_EPICS_BIGQUERY_TABLE_ID", "")

API_DELAY = float(os.getenv("JIRA_API_DELAY", "0.05"))
RETRY_DELAY = 5
MAX_RETRIES = 3
REQUEST_TIMEOUT = int(os.getenv("JIRA_REQUEST_TIMEOUT", "30"))
JIRA_FETCH_WORKERS = int(os.getenv("JIRA_FETCH_WORKERS", "12"))
LLM_TIMEOUT_SECONDS = int(os.getenv("LLM_TIMEOUT_SECONDS", "75"))
LLM_MAX_ATTEMPTS = int(os.getenv("LLM_MAX_ATTEMPTS", "3"))
LLM_RECOVERY_TIMEOUT_SECONDS = int(os.getenv("LLM_RECOVERY_TIMEOUT_SECONDS", "120"))
LLM_PARALLEL_WORKERS = int(os.getenv("LLM_PARALLEL_WORKERS", "8"))
EPIC_PARALLEL_WORKERS = int(os.getenv("EPIC_PARALLEL_WORKERS", "4"))
LLM_STORY_SAMPLE_SIZE = int(os.getenv("LLM_STORY_SAMPLE_SIZE", "6"))
LLM_DESC_MAX_CHARS = int(os.getenv("LLM_DESC_MAX_CHARS", "450"))
LLM_MAX_OUTPUT_TOKENS = int(os.getenv("LLM_MAX_OUTPUT_TOKENS", "8192"))
CAPABILITY_LIMIT = int(os.getenv("CAPABILITY_LIMIT", "0"))


def _read_epic_keys_from_file(file_path):
    keys = []
    with open(file_path, "r", encoding="utf-8") as handle:
        for line in handle:
            stripped = line.strip()
            if stripped and not stripped.startswith("#"):
                keys.append(stripped)
    return keys


def resolve_epic_keys():
    epic_keys_csv = os.getenv("EPIC_KEYS_CSV", "").strip()
    if epic_keys_csv:
        keys = [k.strip() for k in epic_keys_csv.split(",") if k.strip()]
        if keys:
            return keys

    epic_keys_file = os.getenv("EPIC_KEYS_FILE", "").strip()
    if epic_keys_file:
        if not os.path.exists(epic_keys_file):
            raise FileNotFoundError(f"EPIC_KEYS_FILE not found: {epic_keys_file}")
        keys = _read_epic_keys_from_file(epic_keys_file)
        if keys:
            return keys

    # Auto-discover epic_keys.txt next to this script
    default_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "epic_keys.txt")
    if os.path.exists(default_file):
        keys = _read_epic_keys_from_file(default_file)
        if keys:
            return keys

    if DEFAULT_EPIC_KEYS:
        return list(DEFAULT_EPIC_KEYS)

    raise ValueError(
        "No epic keys configured. Set EPIC_KEYS_CSV, EPIC_KEYS_FILE, "
        "or place an epic_keys.txt file next to the script."
    )


EPIC_KEYS = resolve_epic_keys()

# ─── LLM Setup ──────────────────────────────────────────────────────────────
PROJECT_ID = os.getenv("GCP_PROJECT_ID", "")
LOCATION = os.getenv("GCP_LOCATION", "us-central1")
aiplatform.init(project=PROJECT_ID, location=LOCATION)
LLM = VertexAI(
    model_name="gemini-2.5-flash",
    temperature=0.18,
    max_output_tokens=LLM_MAX_OUTPUT_TOKENS,
)

# ═══════════════════════════════════════════════════════════════════════════════
# Jira HTTP Session
# ═══════════════════════════════════════════════════════════════════════════════
_SESSION = None


def _create_session():
    session = requests.Session()
    retry_strategy = Retry(
        total=5,
        backoff_factor=2,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry_strategy, pool_connections=10, pool_maxsize=10)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    if not JIRA_EMAIL or not JIRA_API_TOKEN:
        raise ValueError("Set JIRA_EMAIL and JIRA_API_TOKEN environment variables")
    session.auth = HTTPBasicAuth(JIRA_EMAIL, JIRA_API_TOKEN)
    session.headers.update({"Accept": "application/json", "Content-Type": "application/json"})
    return session


def _get_session():
    global _SESSION
    if _SESSION is None:
        _SESSION = _create_session()
    return _SESSION


def _api_call(method, url, params=None, json_data=None):
    for attempt in range(MAX_RETRIES):
        try:
            session = _get_session()
            resp = session.get(url, params=params, timeout=REQUEST_TIMEOUT) if method == "GET" \
                else session.post(url, json=json_data, timeout=REQUEST_TIMEOUT)
            if resp.status_code == 429:
                wait = int(resp.headers.get("Retry-After", 30))
                logger.warning(f"Rate limited. Waiting {wait}s...")
                time.sleep(wait)
                continue
            time.sleep(API_DELAY)
            return resp
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
            wait = RETRY_DELAY * (attempt + 1)
            logger.warning(f"Connection error (attempt {attempt + 1}/{MAX_RETRIES}): {e}")
            time.sleep(wait)
            global _SESSION
            _SESSION = None
    raise requests.exceptions.ConnectionError(f"Failed after {MAX_RETRIES} retries: {url}")


def jira_get(endpoint, params=None):
    url = f"{JIRA_URL}/rest/api/3/{endpoint}"
    resp = _api_call("GET", url, params=params)
    if not resp.ok:
        logger.error(f"GET {endpoint} → {resp.status_code}: {resp.text[:500]}")
        resp.raise_for_status()
    return resp.json()


def jira_agile_get(endpoint, params=None):
    url = f"{JIRA_URL}/rest/agile/1.0/{endpoint}"
    resp = _api_call("GET", url, params=params)
    if not resp.ok:
        logger.error(f"AGILE GET {endpoint} → {resp.status_code}")
        resp.raise_for_status()
    return resp.json()


# ═══════════════════════════════════════════════════════════════════════════════
# Jira Helpers
# ═══════════════════════════════════════════════════════════════════════════════
_KEY_CACHE = {}


def _normalize_issue(raw):
    norm = {}
    key = raw.get("key", "")
    if key and not str(key).isdigit():
        norm["key"] = key
    elif key:
        norm["key"] = str(key)
    if "id" in raw:
        norm["id"] = str(raw["id"])
    if norm.get("key", "").isdigit():
        flds = raw.get("fields") or {}
        if "key" in flds:
            norm["key"] = flds["key"]
    if "key" not in norm:
        norm["key"] = str(raw.get("id", "UNKNOWN"))
    norm["fields"] = raw.get("fields") or {}
    return norm


def _resolve_key(issue):
    key = issue.get("key", "")
    if key and not key.isdigit():
        return key
    issue_id = issue.get("id", key) or "UNKNOWN"
    if issue_id in _KEY_CACHE:
        issue["key"] = _KEY_CACHE[issue_id]
        return _KEY_CACHE[issue_id]
    try:
        full = jira_get(f"issue/{issue_id}", params={"fields": "summary"})
        real = full.get("key", str(issue_id))
        issue["key"] = real
        _KEY_CACHE[issue_id] = real
        return real
    except Exception:
        return str(issue_id)


def jql_search(jql, fields=None, max_results=500):
    all_issues, start, batch = [], 0, 50
    while True:
        params = {"jql": jql, "startAt": start, "maxResults": batch}
        if fields and fields != "*all":
            params["fields"] = ",".join(fields) if isinstance(fields, list) else fields
        url = f"{JIRA_URL}/rest/api/3/search/jql"
        resp = _api_call("GET", url, params=params)
        if not resp.ok:
            logger.error(f"Search failed: {resp.status_code} - {resp.text[:500]}")
            resp.raise_for_status()
        data = resp.json()
        raw = []
        for k in ("issues", "results", "values", "searchResults"):
            if k in data:
                raw = data[k]
                break
        if not raw and isinstance(data, list):
            raw = data
        issues = [_normalize_issue(i) for i in raw]
        for iss in issues:
            _resolve_key(iss)
        all_issues.extend(issues)
        total = data.get("total", data.get("totalCount", 0))
        if len(raw) < batch or len(all_issues) >= max_results or len(all_issues) >= total:
            break
        start += batch
    return all_issues


# ═══════════════════════════════════════════════════════════════════════════════
# Field Extractors
# ═══════════════════════════════════════════════════════════════════════════════
def _field(issue, name, default=None):
    fields = issue.get("fields") or {}
    return fields.get(name, default)


def _safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val)
        except ValueError:
            return None
    if isinstance(val, dict):
        for k in ("value", "estimate", "originalEstimate", "storyPoints"):
            if k in val:
                return _safe_float(val[k])
        return None
    if isinstance(val, list) and val:
        return _safe_float(val[0])
    return None


def get_story_points(issue):
    fields = issue.get("fields") or {}
    for sp_field in STORY_POINTS_FIELDS:
        val = fields.get(sp_field)
        if val is not None:
            result = _safe_float(val)
            if result is not None:
                return result
    return None


def get_fix_versions(issue):
    fv = _field(issue, "fixVersions", [])
    return [v.get("name", "") for v in fv if isinstance(v, dict)] if fv else []


def is_q3_2026(fix_versions):
    """Check if fix versions indicate Q3 2026 - used for epics & capabilities."""
    for v in fix_versions:
        for p in Q3_FIX_VERSION_PATTERNS:
            if p.lower() in v.lower() and "2026" in v:
                return True
    return False


def is_feature_q3_scoped(feature_fix_versions, capability_fix_versions):
    """Determine if a feature is scoped for Q3."""
    for v in feature_fix_versions:
        if v in PLANNING_SET:
            return True

    feat_candidates = [v for v in feature_fix_versions if "Planning Candidate" in v]
    cap_candidates = [v for v in capability_fix_versions if "Planning Candidate" in v]
    if feat_candidates and cap_candidates:
        for v in feat_candidates:
            if v in cap_candidates:
                return True

    return False


def get_status(issue):
    s = _field(issue, "status", {})
    return s.get("name", "Unknown") if isinstance(s, dict) else (str(s) if s else "Unknown")


def get_summary(issue):
    return _field(issue, "summary", "")


def get_issue_type(issue):
    it = _field(issue, "issuetype", {})
    return it.get("name", "Unknown") if isinstance(it, dict) else str(it)


def get_assignee(issue):
    a = _field(issue, "assignee")
    return a.get("displayName") if isinstance(a, dict) else None


def get_team_name(issue):
    team = _field(issue, TEAM_FIELD)
    if team:
        if isinstance(team, str):
            return team
        if isinstance(team, dict):
            return team.get("name", team.get("value", str(team)))
        if isinstance(team, list) and team:
            first = team[0]
            return first.get("name", str(first)) if isinstance(first, dict) else str(first)
    return ""


def get_team_id(issue):
    team = _field(issue, TEAM_FIELD)
    if not team:
        return ""
    if isinstance(team, dict):
        # Jira team custom field usually carries id/uuid in id.
        return str(team.get("id") or team.get("value") or team.get("name") or "")
    if isinstance(team, list) and team:
        first = team[0]
        if isinstance(first, dict):
            return str(first.get("id") or first.get("value") or first.get("name") or "")
        return str(first)
    return str(team)


def get_project_key(issue):
    project = _field(issue, "project")
    if isinstance(project, dict):
        return project.get("key", "")
    return ""


def get_description_text(issue):
    desc = _field(issue, "description")
    if not desc:
        return ""
    if isinstance(desc, str):
        return desc[:2000]

    def _extract(node):
        texts = []
        if isinstance(node, dict):
            if node.get("type") == "text":
                texts.append(node.get("text", ""))
            for child in node.get("content", []):
                texts.extend(_extract(child))
        elif isinstance(node, list):
            for item in node:
                texts.extend(_extract(item))
        return texts

    return " ".join(_extract(desc))[:2000]


# ═══════════════════════════════════════════════════════════════════════════════
# Jira Data Extraction
# ═══════════════════════════════════════════════════════════════════════════════
_DONE_STATUSES = {"done", "closed", "resolved", "complete", "completed"}
_TEAM_HISTORICAL_CACHE = {}
_TEAM_HISTORICAL_LOCK = threading.Lock()


def _escape_jql_string(value):
    return str(value).replace('\\', '\\\\').replace('"', '\\"')


def fetch_team_historical_metrics(team_name, team_id=None, project_key=None):
    """
    Return (avg_features_per_fix_version, avg_story_points_per_fix_version) for
    completed issues in last 1 year, excluding fix versions containing "Archive".
    """
    if not team_name:
        return "-", "-"

    cache_key = (project_key or "", team_id or "", team_name)
    with _TEAM_HISTORICAL_LOCK:
        cached = _TEAM_HISTORICAL_CACHE.get(cache_key)
    if cached is not None:
        return cached

    project_clause = f"project = {project_key} AND " if project_key else ""
    escaped_team_id = _escape_jql_string(team_id) if team_id else ""
    jql = (
        f"{project_clause}"
        f'issuetype = "Epic (Feature)" AND '
        f"statusCategory = Done AND "
        f'cf[10001] = "{escaped_team_id}" AND '
        f"fixVersion is not EMPTY AND "
        f'fixVersion != "Archive" AND '
        f"resolved >= -365d"
    )
    fields = ["fixVersions"] + STORY_POINTS_FIELDS

    try:
        issues = jql_search(jql, fields=fields, max_results=5000)
        if not issues and team_id:
            # Safe fallback for projects where issue type label differs.
            fallback_jql = (
                f"{project_clause}"
                f"statusCategory = Done AND "
                f'cf[10001] = "{escaped_team_id}" AND '
                f"fixVersion is not EMPTY AND "
                f'fixVersion != "Archive" AND '
                f"resolved >= -365d"
            )
            issues = jql_search(fallback_jql, fields=fields, max_results=5000)
    except Exception as exc:
        logger.warning(f"Could not compute historical metrics for team '{team_name}': {exc}")
        result = ("-", "-")
        with _TEAM_HISTORICAL_LOCK:
            _TEAM_HISTORICAL_CACHE[cache_key] = result
        return result

    by_fix_version = {}

    for issue in issues:
        fix_versions = [
            fv for fv in get_fix_versions(issue)
            if fv and "archive" not in fv.lower()
        ]
        if not fix_versions:
            continue

        story_points = get_story_points(issue) or 0.0
        for fv in set(fix_versions):
            bucket = by_fix_version.setdefault(fv, {"features": 0, "story_points": 0.0})
            bucket["features"] += 1
            bucket["story_points"] += float(story_points)

    if not by_fix_version:
        result = ("0.00", "0.00")
    else:
        fix_version_count = len(by_fix_version)
        total_features = sum(v["features"] for v in by_fix_version.values())
        total_story_points = sum(v["story_points"] for v in by_fix_version.values())
        avg_features = total_features / fix_version_count
        avg_story_points = total_story_points / fix_version_count
        result = (f"{avg_features:.2f}", f"{avg_story_points:.2f}")

    with _TEAM_HISTORICAL_LOCK:
        _TEAM_HISTORICAL_CACHE[cache_key] = result

    return result


def fetch_stories(feature_key):
    logger.debug(f"        Fetching stories for feature {feature_key}")
    issues = jql_search(
        f"parent = {feature_key}",
        fields=["summary", "status", "assignee", "issuetype"] + STORY_POINTS_FIELDS,
    )
    return [
        {
            "key": s.get("key", "UNKNOWN"),
            "summary": get_summary(s),
            "status": get_status(s),
            "story_points": get_story_points(s),
            "assignee": get_assignee(s),
            "issue_type": get_issue_type(s),
        }
        for s in issues
    ]


def fetch_features(capability_key, capability_fix_versions=None):
    feat_fields = [
        "summary", "status", "fixVersions", "description",
        "assignee", "issuetype", TEAM_FIELD, "components", "project",
    ] + STORY_POINTS_FIELDS
    issues = jql_search(f"parent = {capability_key}", fields=feat_fields)
    logger.debug(f"      Capability {capability_key}: {len(issues)} features found")
    cap_fv = capability_fix_versions or []

    feature_base = []
    for idx, f in enumerate(issues):
        summary = get_summary(f)
        description = get_description_text(f)
        fv = get_fix_versions(f)
        feature_base.append({
            "_idx": idx,
            "key": f.get("key", "UNKNOWN"),
            "summary": summary,
            "status": get_status(f),
            "description": description,
            "fix_versions": fv,
            "is_q3_scoped": is_feature_q3_scoped(fv, cap_fv),
            "team_name": get_team_name(f),
            "team_id": get_team_id(f),
            "project_key": get_project_key(f),
            "assignee": get_assignee(f),
            "story_points": get_story_points(f),
        })

    stories_map = {}
    if feature_base:
        logger.debug(
            f"      Capability {capability_key}: fetching stories in parallel "
            f"({JIRA_FETCH_WORKERS} workers)"
        )
        with ThreadPoolExecutor(max_workers=JIRA_FETCH_WORKERS) as executor:
            future_to_key = {executor.submit(fetch_stories, feat["key"]): feat["key"] for feat in feature_base}
            for done_idx, future in enumerate(as_completed(future_to_key), 1):
                f_key = future_to_key[future]
                stories_map[f_key] = future.result()
                if done_idx % 10 == 0 or done_idx == len(feature_base):
                    logger.info(
                        f"      Capability {capability_key}: fetched stories for {done_idx}/{len(feature_base)} features"
                    )

    features = []
    for feat in feature_base:
        f_key = feat["key"]
        stories = stories_map.get(f_key, [])
        total = len(stories)
        done = len([s for s in stories if s["status"].lower() in _DONE_STATUSES])
        total_sp = sum(s["story_points"] or 0 for s in stories)
        done_sp = sum(
            s["story_points"] or 0
            for s in stories
            if s["status"].lower() in _DONE_STATUSES
        )

        if total_sp > 0:
            pct_complete = round(done_sp / total_sp * 100, 1)
        else:
            pct_complete = round(done / total * 100, 1) if total > 0 else 0

        features.append({
            "key": f_key,
            "summary": feat["summary"],
            "status": feat["status"],
            "description": feat["description"],
            "fix_versions": feat["fix_versions"],
            "is_q3_scoped": feat["is_q3_scoped"],
            "team_name": feat["team_name"],
            "team_id": feat.get("team_id", ""),
            "project_key": feat.get("project_key", ""),
            "assignee": feat["assignee"],
            "story_points": feat["story_points"],
            "stories": stories,
            "total_stories": total,
            "done_stories": done,
            "total_story_points": total_sp,
            "done_story_points": done_sp,
            "pct_complete": pct_complete,
        })
    return features


def fetch_capabilities(epic_key):
    cap_fields = ["summary", "status", "fixVersions", "labels", "issuetype", "description"]
    issues = jql_search(f"parent = {epic_key}", fields=cap_fields)
    if CAPABILITY_LIMIT > 0:
        issues = issues[:CAPABILITY_LIMIT]
    logger.debug(f"    Epic {epic_key}: {len(issues)} capabilities found")

    caps = []
    for c in issues:
        cap_summary = get_summary(c)
        cap_description = get_description_text(c)
        c_key = c.get("key", "UNKNOWN")
        fv = get_fix_versions(c)
        features = fetch_features(c_key, capability_fix_versions=fv)
        total = len(features)
        done = len([ft for ft in features if ft["status"].lower() in _DONE_STATUSES])

        caps.append({
            "key": c_key,
            "summary": cap_summary,
            "status": get_status(c),
            "fix_versions": fv,
            "is_q3_picked": is_q3_2026(fv),
            "features": features,
            "total_features": total,
            "done_features": done,
        })
    return caps


def fetch_epic(epic_key):
    """Fetch a single epic with its full hierarchy."""
    try:
        epic = jira_get(f"issue/{epic_key}")
    except requests.exceptions.HTTPError as e:
        logger.error(f"Could not fetch {epic_key}: {e}")
        return None

    epic_summary = get_summary(epic)

    try:
        capabilities = fetch_capabilities(epic_key)
    except Exception as e:
        logger.error(f"Error extracting capabilities for {epic_key}: {e}")
        return None

    total_caps = len(capabilities)
    done_caps = len([c for c in capabilities if c["status"].lower() in _DONE_STATUSES])
    epic_pct = round(done_caps / total_caps * 100) if total_caps > 0 else 0

    return {
        "key": epic_key,
        "summary": epic_summary,
        "status": get_status(epic),
        "fix_versions": get_fix_versions(epic),
        "is_q3": is_q3_2026(get_fix_versions(epic)),
        "epic_pct_completion": epic_pct,
        "total_capabilities": total_caps,
        "done_capabilities": done_caps,
        "capabilities": capabilities,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# LLM Scoring
# ═══════════════════════════════════════════════════════════════════════════════
def score_feature(feature):
    """Score a feature using LLM based on Jira data."""
    stories_summary = json.dumps(
        [{"key": s["key"], "summary": s["summary"], "status": s["status"], "sp": s["story_points"]}
         for s in feature.get("stories", [])[:LLM_STORY_SAMPLE_SIZE]],
        indent=2,
    )

    prompt = f"""You are a software project estimator. Analyze this Jira feature and provide scores.

FEATURE: {feature['key']} - {feature['summary']}
DESCRIPTION: {feature.get('description', 'No description')[:LLM_DESC_MAX_CHARS]}
STATUS: {feature['status']}
TOTAL STORIES: {feature['total_stories']}
DONE STORIES: {feature['done_stories']}
TOTAL STORY POINTS (from Jira): {feature.get('total_story_points', 0)}
FEATURE STORY POINTS: {feature.get('story_points', 'Not set')}
TEAM: {feature.get('team_name', 'Unknown')}

STORIES BREAKDOWN:
{stories_summary}

Based on the REQUIREMENT SCOPE AND COMPLEXITY (not story count), provide:
1. AC Score (%) - How well acceptance criteria are defined (0-100)
2. Story Granularity (%) - How well the feature is decomposed into stories (0-100)
3. Description Granularity (%) - How detailed the feature/story descriptions are (0-100)
4. Sizing Accuracy (%) - How accurate the story point sizing appears (0-100)
5. Test Coverage (%) - Evidence of test stories/tasks (0-100)
6. Overall Quality Score (%) - Weighted average of above (0-100)
7. AI Estimated Story Points - Based on REQUIREMENT SCOPE, not story count. Consider:
   - What is the actual work scope (enterprise-wide vs targeted fix)?
   - Complexity of implementation (compliance, automation, remediation)?
   - Integration points and blast radius?
   - Give a single integer estimate.
8. % Complete wrt Estimate - Your estimate of how complete this feature is relative to its expected scope, based on the Jira evidence provided (0-100).

E2E handling rule (IMPORTANT):
- Keep this item in reporting, but do NOT consider E2E testing-only effort when scoring quality or estimating AI story points.
- If summary/description/stories mention E2E testing, treat that as non-estimation scope and focus on core delivery scope.

Rules for pct_complete (IMPORTANT):
- Do NOT default to 100.
- Return 100 only when evidence strongly supports completion:
    1) feature/status is done-like (Done/Closed/Resolved/Complete), and
    2) stories are mostly/all done, and
    3) remaining work appears negligible.
- If feature is active/in progress, pct_complete should usually be below 90.
- If status is Won't Do/Cancelled/Rejected, pct_complete should usually be 0-20.
- Use done stories, story statuses, and description context as evidence; avoid optimistic bias.
- Prefer realistic mid-range values (30-85) when evidence is mixed.

Respond ONLY in this exact JSON format, no other text:
{{"ac_score": <int>, "story_granularity": <int>, "desc_granularity": <int>, "sizing_accuracy": <int>, "test_coverage": <int>, "overall_quality": <int>, "ai_estimated_sp": <int>, "pct_complete": <int>}}"""

    def _invoke_llm_with_timeout(prompt_text, timeout_seconds):
        result_queue = queue.Queue(maxsize=1)

        def _worker():
            try:
                result_queue.put(("ok", LLM.invoke(prompt_text)))
            except Exception as ex:
                result_queue.put(("err", ex))

        thread = threading.Thread(target=_worker, daemon=True)
        thread.start()
        thread.join(timeout_seconds)

        if thread.is_alive():
            raise FutureTimeoutError()

        status, payload = result_queue.get_nowait()
        if status == "err":
            raise payload
        return payload

    last_error = None
    for attempt in range(1, LLM_MAX_ATTEMPTS + 1):
        try:
            response = _invoke_llm_with_timeout(prompt, LLM_TIMEOUT_SECONDS)
            response = str(response).strip()
            # Robust JSON extraction: find first { and last }
            start_idx = response.find("{")
            end_idx = response.rfind("}")
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                response = response[start_idx:end_idx + 1]
            result = json.loads(response)
            logger.info(
                f"        OK {feature['key']}: Quality={result.get('overall_quality')}%, "
                f"AI SP={result.get('ai_estimated_sp')}, AC={result.get('ac_score')}%, "
                f"Complete={result.get('pct_complete')}%"
            )
            return result
        except FutureTimeoutError as e:
            last_error = e
            logger.warning(
                f"        TIMEOUT LLM for {feature['key']} "
                f"(attempt {attempt}/{LLM_MAX_ATTEMPTS}, {LLM_TIMEOUT_SECONDS}s) - retrying..."
            )
            time.sleep(2 * attempt)  # backoff: 2s, 4s, 6s...
        except Exception as e:
            last_error = e
            logger.warning(
                f"        WARN LLM error for {feature['key']} "
                f"(attempt {attempt}/{LLM_MAX_ATTEMPTS}): {e} - retrying..."
            )
            time.sleep(2 * attempt)

    # All attempts exhausted - raise so caller can decide
    raise RuntimeError(
        f"LLM scoring failed for {feature['key']} after {LLM_MAX_ATTEMPTS} attempts: {last_error}"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════
def _fallback_scores_for_feature(feature):
    """Return conservative fallback scores when LLM scoring fails."""
    return {
        "ac_score": 0,
        "story_granularity": 0,
        "desc_granularity": 0,
        "sizing_accuracy": 0,
        "test_coverage": 0,
        "overall_quality": 0,
        "ai_estimated_sp": int(feature.get("story_points") or 0),
        "pct_complete": int(feature.get("pct_complete", 0)),
    }


def process_single_epic(epic_key, start_time):
    """Fetch, score, and flatten one epic for reporting."""
    logger.info(f"Processing Epic: {epic_key}")

    epic_data = fetch_epic(epic_key)
    if epic_data is None:
        logger.warning(f"Skipped {epic_key} (fetch failed)")
        return {
            "epic_key": epic_key,
            "ok": False,
            "epic_data": None,
            "rows": [],
            "skipped_features": [],
        }

    elapsed = time.time() - start_time
    logger.info(
        f"OK {epic_key}: {epic_data['epic_pct_completion']}% done | "
        f"Elapsed: {elapsed/60:.1f} min"
    )

    epic_features = [
        feat
        for cap in epic_data["capabilities"]
        for feat in cap["features"]
    ]
    total_feats = len(epic_features)
    logger.info(f"INFO {epic_key}: scoring {total_feats} features ({LLM_PARALLEL_WORKERS} workers)")

    scored_map = {}
    skipped_features = []
    completed = 0
    with ThreadPoolExecutor(max_workers=LLM_PARALLEL_WORKERS) as executor:
        future_to_feat = {executor.submit(score_feature, feat): feat for feat in epic_features}
        for future in as_completed(future_to_feat):
            feat = future_to_feat[future]
            completed += 1
            try:
                scored_map[feat["key"]] = future.result()
            except Exception as e:
                logger.warning(f"WARN Initial scoring failed for {feat['key']}: {e}")
                original_timeout = LLM_TIMEOUT_SECONDS
                try:
                    globals()["LLM_TIMEOUT_SECONDS"] = max(original_timeout, LLM_RECOVERY_TIMEOUT_SECONDS)
                    scored_map[feat["key"]] = score_feature(feat)
                    logger.info(f"OK Recovery scoring succeeded for {feat['key']}")
                except Exception as recovery_error:
                    logger.error(f"FAIL Using fallback scores for {feat['key']}: {recovery_error}")
                    skipped_features.append(feat["key"])
                    scored_map[feat["key"]] = _fallback_scores_for_feature(feat)
                finally:
                    globals()["LLM_TIMEOUT_SECONDS"] = original_timeout
            if completed % 10 == 0 or completed == total_feats:
                elapsed_now = time.time() - start_time
                logger.info(f"{epic_key}: scored {completed}/{total_feats} | {elapsed_now/60:.1f} min elapsed")

    epic_rows = build_excel_rows(epic_data, scored_map)

    if SAVE_PER_EPIC_EXCEL:
        per_epic_file = f"{epic_key}_Quality_Report.xlsx"
        export_to_excel(epic_rows, per_epic_file)
        logger.info(f"SAVED per-epic report: {per_epic_file}")

    return {
        "epic_key": epic_key,
        "ok": True,
        "epic_data": epic_data,
        "rows": epic_rows,
        "skipped_features": skipped_features,
    }


def main():
    start_time = time.time()
    logger.info(f"Enhanced Jira Quality Report - {len(EPIC_KEYS)} epics")
    logger.info(f"Jira URL: {JIRA_URL}")
    logger.info(f"Epic parallel workers: {EPIC_PARALLEL_WORKERS}")
    logger.info(f"Capability limit: {CAPABILITY_LIMIT if CAPABILITY_LIMIT > 0 else 'No limit'}")
    logger.info(f"BigQuery target table: {ALL_EPICS_BIGQUERY_TABLE_ID}")
    if ENABLE_BIGQUERY_UPLOAD:
        logger.info(f"BigQuery mode: {'APPEND_PER_EPIC' if BQ_APPEND_PER_EPIC else 'BULK_REFRESH'}")

    all_rows = []
    success = 0
    skipped_features = []

    if ENABLE_BIGQUERY_UPLOAD and BQ_APPEND_PER_EPIC and BQ_CLEAR_TABLE_BEFORE_RUN:
        clear_bigquery_table(ALL_EPICS_BIGQUERY_TABLE_ID)

    with ThreadPoolExecutor(max_workers=EPIC_PARALLEL_WORKERS) as executor:
        future_to_epic = {executor.submit(process_single_epic, epic_key, start_time): epic_key for epic_key in EPIC_KEYS}
        for done_count, future in enumerate(as_completed(future_to_epic), 1):
            epic_key = future_to_epic[future]
            try:
                result = future.result()
            except Exception as e:
                logger.error(f"FAIL {epic_key}: unexpected error in epic worker: {e}")
                continue

            if not result["ok"]:
                continue

            success += 1
            all_rows.extend(result["rows"])
            skipped_features.extend(result["skipped_features"])
            if ENABLE_BIGQUERY_UPLOAD and BQ_APPEND_PER_EPIC and result["rows"]:
                upload_all_epic_report_to_bigquery(
                    result["rows"],
                    ALL_EPICS_BIGQUERY_TABLE_ID,
                    write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
                )
            logger.info(f"Progress: epics completed {done_count}/{len(EPIC_KEYS)}")

    # ── Final outputs ────────────────────────────────────────────────────────
    if ENABLE_BIGQUERY_UPLOAD and not BQ_APPEND_PER_EPIC:
        upload_all_epic_report_to_bigquery(all_rows, ALL_EPICS_BIGQUERY_TABLE_ID)

    if SAVE_COMBINED_EXCEL:
        export_to_excel(all_rows, OUTPUT_FILE_EXCEL)

    elapsed = time.time() - start_time
    logger.info("=" * 70)
    if ENABLE_BIGQUERY_UPLOAD:
        logger.info(f"OK BigQuery table refreshed: {ALL_EPICS_BIGQUERY_TABLE_ID}")
    if SAVE_COMBINED_EXCEL:
        logger.info(f"OK Excel Report saved: {OUTPUT_FILE_EXCEL}")
    logger.info(f"INFO Epics: {success}/{len(EPIC_KEYS)} processed")
    logger.info(f"INFO All_epic_report rows: {len(all_rows)}")
    logger.info(f"INFO Skipped Features: {len(skipped_features)}")
    if skipped_features:
        logger.info(f"Skipped feature keys: {', '.join(skipped_features[:25])}")
    logger.info(f"Total time: {elapsed/60:.1f} min")
    logger.info("=" * 70)


# ═══════════════════════════════════════════════════════════════════════════════
# Excel Styles and Functions
# ═══════════════════════════════════════════════════════════════════════════════
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
NO_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
YES_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
DONE_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

EXCEL_HEADERS = [
    "Epic Name & ID", "Epic % Completion in Jira", "Capability",
    "Capability Picked for Q3-2026", "Capability Coverage wrt Epic %",
    "Capability Status", "Team Name", "Team Historical # of features / release",
    "Team Historical story Points / release", "Feature Name",
    "Feature Scoped for Q3", "Feature Coverage wrt Capability%",
    "Jira Story points", "No. of Stories Created", "Jira Status",
    "AC Score", "Story Granularity", "Description Granularity",
    "Sizing Accuracy", "Test Coverage", "Overall Quality Score",
    "AI estimated Story Point", "% Complete wrt Estimate", "Team Contact (PO)",
]

EXCEL_ROW_KEYS = [
    "epic_key", "epic_pct", "cap_key", "cap_q3", "cap_coverage",
    "cap_status", "team", "hist_features", "hist_sp", "feat_key",
    "feat_q3", "feat_coverage", "jira_sp", "num_stories", "jira_status",
    "ac", "story_gran", "desc_gran", "sizing_acc", "test_cov",
    "overall_q", "ai_sp", "pct_complete", "po",
]

def build_excel_rows(epic_data, scored_map=None):
    """Build flat rows for one epic using pre-computed scored_map (or inline scoring as fallback)."""
    rows = []
    epic_key = epic_data["key"]
    epic_name = epic_data.get("summary", "")
    epic_pct = f"{epic_data['epic_pct_completion']}%"

    if not epic_data["capabilities"]:
        return rows

    total_features_in_epic = sum(len(c["features"]) for c in epic_data["capabilities"])

    for cap in epic_data["capabilities"]:
        cap_key = cap["key"]
        cap_status = cap["status"]
        cap_is_q3 = cap["is_q3_picked"]
        total_stories_in_cap = sum(f["total_stories"] for f in cap["features"])
        cap_coverage = round(len(cap["features"]) / total_features_in_epic * 100) if total_features_in_epic > 0 else 0

        cap_q3_label = "Yes" if cap_is_q3 else "No"

        if not cap["features"]:
            continue

        for feat in cap["features"]:
            feat_coverage = round(feat["total_stories"] / total_stories_in_cap * 100) if total_stories_in_cap > 0 else 0
            team_name = feat.get("team_name", "Unknown")
            hist_features, hist_sp = fetch_team_historical_metrics(
                team_name,
                team_id=feat.get("team_id"),
                project_key=feat.get("project_key"),
            )

            # Use pre-scored map if available, else score inline
            if scored_map is not None:
                scores = scored_map[feat["key"]]  # KeyError if missing = intentional, run failed
            else:
                logger.info(f"  Scoring {feat['key']}: {feat['summary'][:50]}...")
                scores = score_feature(feat)  # raises if LLM fails all retries
            time.sleep(0)
            
            # Extract quality metrics from LLM response
            ai_sp = scores["ai_estimated_sp"]
            ac = f"{scores.get('ac_score', 0)}%"
            story_gran = f"{scores.get('story_granularity', 0)}%"
            desc_gran = f"{scores.get('desc_granularity', 0)}%"
            sizing_acc = f"{scores.get('sizing_accuracy', 0)}%"
            test_cov = f"{scores.get('test_coverage', 0)}%"
            overall_q = f"{scores.get('overall_quality', 0)}%"
            pct_complete = f"{scores.get('pct_complete', feat.get('pct_complete', 0))}%"
            po = feat.get("assignee") or "-"

            rows.append({
                "epic_key": epic_key,
                "epic_name": epic_name,
                "epic_pct": epic_pct,
                "cap_key": cap_key,
                "cap_q3": cap_q3_label,
                "cap_coverage": f"{cap_coverage}%",
                "cap_status": cap_status,
                "team": team_name,
                "hist_features": hist_features,
                "hist_sp": hist_sp,
                "feat_key": feat["key"],
                "feat_q3": "Yes" if feat["is_q3_scoped"] else "No",
                "feat_coverage": f"{feat_coverage}%",
                "jira_sp": feat.get("total_story_points") if feat.get("total_stories", 0) > 0 else (feat.get("story_points") or "-"),
                "num_stories": feat["total_stories"],
                "jira_status": feat["status"],
                "ac": ac,
                "story_gran": story_gran,
                "desc_gran": desc_gran,
                "sizing_acc": sizing_acc,
                "test_cov": test_cov,
                "overall_q": overall_q,
                "ai_sp": ai_sp,
                "pct_complete": pct_complete,
                "po": po,
            })

    return rows


def write_excel_detail_headers(ws):
    """Write header row to Excel sheet."""
    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def append_excel_detail_rows(ws, rows, start_row):
    """Append rows to the detail sheet."""
    for r_idx, row_data in enumerate(rows, start_row):
        for c_idx, key in enumerate(EXCEL_ROW_KEYS, 1):
            val = row_data.get(key, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER
            
            # Color code by overall quality score
            if key == "overall_q":
                score_val = str(val).replace("%", "")
                try:
                    score_num = int(score_val)
                    if score_num >= 60:
                        cell.fill = YES_FILL
                    elif score_num >= 30:
                        cell.fill = NO_FILL
                    else:
                        cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
                except ValueError:
                    pass
    
    return start_row + len(rows)


def set_excel_detail_widths(ws):
    """Set column widths for detail sheet."""
    col_widths = [20, 12, 18, 12, 12, 12, 25, 12, 12, 25, 10, 12, 10, 10, 14,
                  8, 10, 10, 10, 10, 10, 10, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_to_excel(all_rows, output_file):
    """Export rows to Excel detail sheet only."""
    import time
    
    wb = Workbook()
    
    # Detail sheet
    ws_detail = wb.active
    ws_detail.title = "Quality Scores Detail"
    write_excel_detail_headers(ws_detail)
    
    append_excel_detail_rows(ws_detail, all_rows, 2)
    set_excel_detail_widths(ws_detail)
    
    # Try to remove old file with retry
    max_retries = 5
    for attempt in range(max_retries):
        try:
            if os.path.exists(output_file):
                os.remove(output_file)
            break
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(0.5)
                logger.debug(f"Retry {attempt + 1}: Could not remove old file, retrying...")
            else:
                logger.debug(f"Could not remove old file after {max_retries} attempts: {e}")
    
    # Save
    try:
        wb.save(output_file)
        logger.info(f"OK Excel Report saved: {output_file}")
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        raise


def clear_bigquery_table(table_id):
    """Clear target table once at the start of a run."""
    client = bigquery.Client(project=PROJECT_ID)
    logger.info(f"Clearing BigQuery table before run: {table_id}")
    client.query(f"TRUNCATE TABLE `{table_id}`").result()


def upload_all_epic_report_to_bigquery(all_rows, table_id, write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE):
    """Load all_epic_report rows to BigQuery with chosen write mode."""
    client = bigquery.Client(project=PROJECT_ID)

    if not all_rows:
        logger.warning("No all_epic_report rows generated in this run; skipping BigQuery refresh.")
        return

    def _to_float(value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        if not text or text == "-":
            return None
        text = text.replace("%", "").replace(",", "")
        try:
            return float(text)
        except ValueError:
            return None

    def _to_int(value):
        as_float = _to_float(value)
        return int(as_float) if as_float is not None else None

    def _to_bool(value):
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in {"yes", "true", "1"}:
            return True
        if text in {"no", "false", "0"}:
            return False
        return None

    def _to_string(value):
        if value is None:
            return None
        text = str(value).strip()
        return text if text and text != "-" else None

    payload = []
    for row in all_rows:
        payload.append({
            "epic": _to_string(row.get("epic_key")),
            "epic_name": _to_string(row.get("epic_name", row.get("epic_key"))),
            "epic_percent_completion_in_jira": _to_float(row.get("epic_pct")),
            "capability": _to_string(row.get("cap_key")),
            "capability_picked_for_q3_2026": _to_bool(row.get("cap_q3")),
            "capability_coverage_wrt_epic_percent": _to_float(row.get("cap_coverage")),
            "capability_status": _to_string(row.get("cap_status")),
            "team_name": _to_string(row.get("team")),
            "team_historical_number_of_features_release": _to_string(row.get("hist_features")),
            "team_historical_story_points_release": _to_string(row.get("hist_sp")),
            "feature_name": _to_string(row.get("feat_key")),
            "feature_scoped_for_q3": _to_bool(row.get("feat_q3")),
            "feature_coverage_wrt_capability_percent": _to_float(row.get("feat_coverage")),
            "jira_story_points": _to_string(row.get("jira_sp")),
            "no_of_stories_created": _to_int(row.get("num_stories")),
            "jira_status": _to_string(row.get("jira_status")),
            "ac_score": _to_float(row.get("ac")),
            "story_granularity": _to_float(row.get("story_gran")),
            "description_granularity": _to_float(row.get("desc_gran")),
            "sizing_accuracy": _to_float(row.get("sizing_acc")),
            "test_coverage": _to_float(row.get("test_cov")),
            "overall_quality_score": _to_float(row.get("overall_q")),
            "ai_estimated_story_point": _to_int(row.get("ai_sp")),
            "percent_complete_wrt_estimate": _to_float(row.get("pct_complete")),
            "team_contact_po": _to_string(row.get("po")),
        })

    logger.info(f"Loading {len(payload)} rows to BigQuery ({write_disposition}): {table_id}")
    bq_schema = [
        bigquery.SchemaField("epic", "STRING"),
        bigquery.SchemaField("epic_name", "STRING"),
        bigquery.SchemaField("epic_percent_completion_in_jira", "FLOAT64"),
        bigquery.SchemaField("capability", "STRING"),
        bigquery.SchemaField("capability_picked_for_q3_2026", "BOOL"),
        bigquery.SchemaField("capability_coverage_wrt_epic_percent", "FLOAT64"),
        bigquery.SchemaField("capability_status", "STRING"),
        bigquery.SchemaField("team_name", "STRING"),
        bigquery.SchemaField("team_historical_number_of_features_release", "STRING"),
        bigquery.SchemaField("team_historical_story_points_release", "STRING"),
        bigquery.SchemaField("feature_name", "STRING"),
        bigquery.SchemaField("feature_scoped_for_q3", "BOOL"),
        bigquery.SchemaField("feature_coverage_wrt_capability_percent", "FLOAT64"),
        bigquery.SchemaField("jira_story_points", "STRING"),
        bigquery.SchemaField("no_of_stories_created", "INT64"),
        bigquery.SchemaField("jira_status", "STRING"),
        bigquery.SchemaField("ac_score", "FLOAT64"),
        bigquery.SchemaField("story_granularity", "FLOAT64"),
        bigquery.SchemaField("description_granularity", "FLOAT64"),
        bigquery.SchemaField("sizing_accuracy", "FLOAT64"),
        bigquery.SchemaField("test_coverage", "FLOAT64"),
        bigquery.SchemaField("overall_quality_score", "FLOAT64"),
        bigquery.SchemaField("ai_estimated_story_point", "INT64"),
        bigquery.SchemaField("percent_complete_wrt_estimate", "FLOAT64"),
        bigquery.SchemaField("team_contact_po", "STRING"),
    ]

    load_config = bigquery.LoadJobConfig(
        source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
        write_disposition=write_disposition,
        schema=bq_schema,
    )
    load_job = client.load_table_from_json(payload, table_id, job_config=load_config)
    load_job.result()

    table_obj = client.get_table(table_id)
    logger.info(f"OK BigQuery all_epic_report refreshed: {table_id} ({table_obj.num_rows} rows)")


if __name__ == "__main__":
    main()
